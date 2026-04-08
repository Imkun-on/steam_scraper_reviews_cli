"""Rich CLI for Steam Reviews Scraper."""

import csv
import os
import sys
from datetime import datetime

# Force UTF-8 output on Windows
if sys.platform == "win32":
    os.environ.setdefault("PYTHONIOENCODING", "utf-8")
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8")

from rich.align import Align
from rich.box import DOUBLE, ROUNDED, SIMPLE_HEAVY
from rich.console import Console, Group
from rich.live import Live
from rich.panel import Panel
from rich.progress import (
    Progress, SpinnerColumn, BarColumn, TextColumn,
    MofNCompleteColumn, TimeElapsedColumn, TimeRemainingColumn,
    TaskProgressColumn,
)
from rich.prompt import Prompt, IntPrompt, Confirm
from rich.rule import Rule
from rich.style import Style
from rich.table import Table
from rich.text import Text

from api import (
    search_games, enrich_search_results, get_app_details, parse_metadata,
    get_review_summary, get_review_counts_by_language, fetch_reviews,
    browse_category, STEAM_LANGUAGES, BROWSE_CATEGORIES,
)
from specs import get_pc_specs

console = Console()

# ─── Symbols ──────────────────────────────────────────────────────────────────
SYM_OK = "[bold green]✓[/bold green]"
SYM_FAIL = "[bold red]✗[/bold red]"
SYM_ARROW = "[bright_cyan]→[/bright_cyan]"
SYM_DOT = "[dim]·[/dim]"


# ─── Banner ───────────────────────────────────────────────────────────────────

def _print_banner() -> None:
    banner_lines = [
        r" ______     ______   ______     ______     __    __        ______     ______     __   __   __     ______     __     __     ______    ",
        r'/\  ___\   /\__  _\ /\  ___\   /\  __ \   /\ "-./  \      /\  == \   /\  ___\   /\ \ / /  /\ \   /\  ___\   /\ \  _ \ \   /\  ___\   ',
        r"\ \___  \  \/_/\ \/ \ \  __\   \ \  __ \  \ \ \-./\ \     \ \  __<   \ \  __\   \ \ \'/   \ \ \  \ \  __\   \ \ \/ " + r'".\ \  \ \___  \  ',
        r' \/\_____\    \ \_\  \ \_____\  \ \_\ \_\  \ \_\ \ \_\     \ \_\ \_\  \ \_____\  \ \__|    \ \_\  \ \_____\  \ \__/".~\_\  \/\_____\ ',
        r"  \/_____/     \/_/   \/_____/   \/_/\/_/   \/_/  \/_/      \/_/ /_/   \/_____/   \/_/      \/_/   \/_____/   \/_/   \/_/   \/_____/ ",
    ]
    colors = [
        "bright_magenta", "magenta", "bright_blue",
        "bright_cyan", "blue",
    ]
    text = Text()
    for i, line in enumerate(banner_lines):
        text.append(line + "\n", style=Style(color=colors[i % len(colors)], bold=True))

    console.print()
    console.print(Panel(
        Align.center(text),
        border_style="bright_blue",
        box=DOUBLE,
        padding=(1, 2),
        expand=False,
    ))


# ─── Search Results ──────────────────────────────────────────────────────────

def show_search_results(results: list[dict], title: str = "Search Results") -> int | None:
    """Display search results and let user pick one, or press q to go back."""
    if not results:
        console.print(f"  {SYM_FAIL} [bold red]No results found.[/bold red]")
        return None

    table = Table(
        title=f"{title} ({len(results)})",
        border_style="bright_cyan",
        box=ROUNDED,
        title_style="bold bright_cyan",
        row_styles=["", "dim"],
    )
    table.add_column("#", style="bold bright_white", justify="center", width=4)
    table.add_column("Title", style="bold white", min_width=25)
    table.add_column("App ID", style="bright_magenta", justify="center", width=10)
    table.add_column("Release Date", style="bright_cyan", justify="center", width=14)
    table.add_column("Languages", style="dim", max_width=30, no_wrap=False)
    table.add_column("Price", style="bold bright_yellow", justify="right", width=14)

    for i, game in enumerate(results, 1):
        table.add_row(
            str(i),
            game["name"],
            str(game["appid"]),
            game.get("release_date", "N/A"),
            game.get("languages", "N/A"),
            game.get("price", "N/A"),
        )

    console.print(table)
    console.print(f"  [dim]Enter [bold]q[/bold] to go back[/dim]")
    console.print()

    valid = [str(i) for i in range(1, len(results) + 1)] + ["q"]
    choice = Prompt.ask(
        "[bold]Select a game (number) or q to go back[/bold]",
        choices=valid,
    )
    if choice == "q":
        return None
    return results[int(choice) - 1]["appid"]


# ─── Main Menu ───────────────────────────────────────────────────────────────

def show_main_menu() -> str:
    """Show the main menu and return the user's choice."""
    categories = list(BROWSE_CATEGORIES.items())

    table = Table(
        title="Main Menu",
        border_style="bright_cyan",
        box=ROUNDED,
        title_style="bold bright_cyan",
    )
    table.add_column("#", style="bold bright_white", justify="center", width=4)
    table.add_column("Option", style="bold white")

    table.add_row("[bright_yellow]S[/bright_yellow]", "[bold]Search by name[/bold]")
    for i, (key, label) in enumerate(categories, 1):
        table.add_row(str(i), label)
    table.add_row("[bright_red]Q[/bright_red]", "[dim]Quit[/dim]")

    console.print()
    console.print(table)
    console.print()

    valid = ["s"] + [str(i) for i in range(1, len(categories) + 1)] + ["q"]
    choice = Prompt.ask(
        "[bold]Choose an option[/bold]",
        choices=valid,
    )

    if choice == "q":
        return "quit"
    if choice == "s":
        return "search"
    idx = int(choice) - 1
    return categories[idx][0]  # return the category key


# ─── Game Details ─────────────────────────────────────────────────────────────

def _format_description(desc: str, width: int = 80) -> list[str]:
    """Format a Steam description into structured Rich markup lines."""
    import textwrap

    lines = []
    raw_lines = desc.split("\n")

    for raw_line in raw_lines:
        stripped = raw_line.strip()
        if not stripped:
            lines.append("")
            continue

        # Bullet points (lines starting with -)
        if stripped.startswith("- "):
            wrapped = textwrap.wrap(stripped[2:], width=width - 6)
            lines.append(f"    [bright_yellow]•[/bright_yellow] [dim]{wrapped[0]}[/dim]")
            for cont in wrapped[1:]:
                lines.append(f"      [dim]{cont}[/dim]")
        # Section headers: short lines (< 60 chars) that are titles/labels
        elif len(stripped) < 60 and (stripped[-1] == ":" or stripped[-1] not in ".,:;!?)\"'") and not stripped.startswith("Notes"):
            lines.append("")
            lines.append(f"  [bold bright_cyan]{stripped}[/bold bright_cyan]")
        # Notes line
        elif stripped.startswith("Notes"):
            lines.append("")
            lines.append(f"  [bold bright_yellow]{stripped}[/bold bright_yellow]")
        else:
            wrapped = textwrap.wrap(stripped, width=width)
            for wl in wrapped:
                lines.append(f"  [dim]{wl}[/dim]")

    # Remove leading/trailing empty lines
    while lines and lines[0] == "":
        lines.pop(0)
    while lines and lines[-1] == "":
        lines.pop()

    return lines


def _parse_package_line(text: str) -> tuple[str, str, str]:
    """Parse a package string like 'Name - 39,99€ 15,99€' into (name, original, discounted)."""
    import re
    # Match patterns like "39,99€ 15,99€" or "39,99€"
    prices = re.findall(r'[\d]+[.,]\d+\s*€', text)
    # Remove prices from the name
    name = re.sub(r'\s*-?\s*[\d]+[.,]\d+\s*€.*', '', text).strip()
    if not name:
        name = text.strip()

    if len(prices) >= 2:
        return name, prices[0].strip(), prices[1].strip()
    elif len(prices) == 1:
        return name, prices[0].strip(), ""
    return name, "", ""


def show_metadata(meta: dict, review_summary: dict):
    # --- Info lines ---
    details = []
    details.append(f"  {SYM_DOT} Release date: [bright_cyan]{meta['release_date']}[/bright_cyan]")
    details.append(f"  {SYM_DOT} Developer: [bright_cyan]{meta['developer']}[/bright_cyan]")
    details.append(f"  {SYM_DOT} Publisher: [bright_cyan]{meta['publisher']}[/bright_cyan]")
    details.append(f"  {SYM_DOT} Tags: [bright_magenta]{', '.join(meta['tags']) if meta['tags'] else 'N/A'}[/bright_magenta]")
    details.append(f"  {SYM_DOT} Franchise: [dim]{meta['franchise']}[/dim]")
    details.append(f"  {SYM_DOT} Price: [bold bright_yellow]{meta['price']}[/bold bright_yellow]")

    # --- Packages table ---
    if meta["packages"]:
        details.append("")
        pkg_table_lines = []
        has_discount = any(_parse_package_line(p)[2] for p in meta["packages"][:5])

        for pkg in meta["packages"][:5]:
            name, original, discounted = _parse_package_line(pkg)
            if discounted:
                pkg_table_lines.append(
                    f"      [bright_yellow]▸[/bright_yellow] {name}  "
                    f"[dim strikethrough]{original}[/dim strikethrough]  "
                    f"[bold bright_green]{discounted}[/bold bright_green]"
                )
            elif original:
                pkg_table_lines.append(
                    f"      [bright_yellow]▸[/bright_yellow] {name}  "
                    f"[bold bright_yellow]{original}[/bold bright_yellow]"
                )
            else:
                pkg_table_lines.append(f"      [bright_yellow]▸[/bright_yellow] [dim]{pkg}[/dim]")

        details.extend(pkg_table_lines)

    # --- Reviews ---
    total = review_summary["total"]
    pos = review_summary["positive"]
    neg = review_summary["negative"]
    pct = (pos / total * 100) if total else 0
    details.append("")
    details.append(
        f"  {SYM_DOT} Reviews: [bold]{total:,}[/bold] total  "
        f"([bold green]{pos:,} positive[/bold green] / [bold red]{neg:,} negative[/bold red])  "
        f"— [bright_yellow]{pct:.0f}%[/bright_yellow]"
    )

    # --- Description ---
    details.append("")
    details.append(f"  [bold bright_cyan]Description[/bold bright_cyan]")
    details.append(f"  [dim]{'─' * 70}[/dim]")
    desc_lines = _format_description(meta["description"])
    details.extend(desc_lines)

    # --- System Requirements ---
    console.print(f"\n  {SYM_ARROW} [bold bright_cyan]Detecting PC specs...[/bold bright_cyan]")
    specs = get_pc_specs()

    details.append("")
    details.append(f"  [bold bright_cyan]System Requirements vs Your PC[/bold bright_cyan]")
    details.append(f"  [dim]{'─' * 70}[/dim]")

    req_table = Table(
        box=None, show_header=True, padding=(0, 2), expand=False,
        header_style="bold bright_white",
    )
    req_table.add_column("Your PC", style="bold green", width=40)
    req_table.add_column("Minimum", style="bright_yellow", width=35)
    req_table.add_column("Recommended", style="white", width=35)

    pc_info = (
        f"[bold]OS:[/bold] {specs['os']}\n"
        f"[bold]CPU:[/bold] {specs['cpu']}\n"
        f"[bold]GPU:[/bold] {specs['gpu']}\n"
        f"[bold]RAM:[/bold] {specs['ram']}"
    )

    req_table.add_row(
        pc_info,
        meta["requirements_min"] if meta["requirements_min"] != "N/A" else "[dim]Not available[/dim]",
        meta["requirements_rec"] if meta["requirements_rec"] != "N/A" else "[dim]Not available[/dim]",
    )

    # --- Main panel ---
    console.print()
    console.print(Panel(
        Group(
            Text.from_markup("\n".join(details)),
            Text(),
            req_table,
        ),
        title=f"[bold bright_white]{meta['title']}[/bold bright_white]",
        border_style="bright_cyan",
        box=DOUBLE,
        expand=False,
        padding=(1, 3),
    ))


# ─── Language Selection ───────────────────────────────────────────────────────

def choose_language(appid: int) -> str:
    lang_codes = list(STEAM_LANGUAGES.keys())

    with console.status("[bright_cyan]Fetching review counts per language...[/bright_cyan]", spinner="dots"):
        counts = get_review_counts_by_language(appid)

    # Color mapping for review sentiments
    _SCORE_STYLES = {
        "Overwhelmingly Positive": "bold bright_green",
        "Very Positive": "bold green",
        "Positive": "green",
        "Mostly Positive": "green",
        "Mixed": "bright_yellow",
        "Mostly Negative": "red",
        "Negative": "red",
        "Very Negative": "bold red",
        "Overwhelmingly Negative": "bold bright_red",
    }

    table = Table(
        title=f"Language Filter ({len(lang_codes)})",
        border_style="bright_cyan",
        box=ROUNDED,
        title_style="bold bright_cyan",
        row_styles=["", "dim"],
    )
    table.add_column("#", style="bold bright_white", justify="center", width=4)
    table.add_column("Code", style="bright_magenta", width=12)
    table.add_column("Language", style="white", min_width=25)
    table.add_column("Reviews", style="bold bright_yellow", justify="right", width=10)
    table.add_column("Rating", justify="center", min_width=28)

    for i, code in enumerate(lang_codes, 1):
        info = counts.get(code, {"total": 0, "desc": "N/A"})
        total = info["total"]
        desc = info["desc"]

        count_str = f"{total:,}" if total > 0 else "[dim]0[/dim]"

        style = _SCORE_STYLES.get(desc, "dim")
        rating_str = f"[{style}]{desc}[/{style}]" if total > 0 else "[dim]N/A[/dim]"

        table.add_row(str(i), code, STEAM_LANGUAGES[code], count_str, rating_str)

    console.print(table)
    console.print()

    choice = IntPrompt.ask(
        "[bold]Select language (number)[/bold]",
        choices=[str(i) for i in range(1, len(lang_codes) + 1)],
    )
    return lang_codes[choice - 1]


# ─── Review Scraping ─────────────────────────────────────────────────────────

def scrape_reviews(appid: int, game_name: str):
    console.print()
    console.rule("[bold bright_blue]Review Scraping[/bold bright_blue]", style="bright_blue")
    console.print()

    lang = choose_language(appid)
    lang_label = STEAM_LANGUAGES[lang]
    console.print(f"\n  {SYM_ARROW} Filter: [bold bright_cyan]{lang_label}[/bold bright_cyan]\n")

    with Progress(
        SpinnerColumn("dots", style="bright_blue"),
        TextColumn("[bold bright_blue]{task.description}"),
        BarColumn(bar_width=50, style="bar.back", complete_style="bright_blue", finished_style="bright_green"),
        TaskProgressColumn(),
        MofNCompleteColumn(),
        TextColumn("[dim]│[/dim]"),
        TimeElapsedColumn(),
        TextColumn("[dim]→[/dim]"),
        TimeRemainingColumn(),
        console=console,
        expand=False,
    ) as progress:
        task = progress.add_task("Downloading reviews", total=None)

        def on_batch(fetched, total):
            progress.update(task, total=total, completed=fetched)

        reviews = fetch_reviews(appid, language=lang, on_batch=on_batch)
        progress.update(task, completed=progress.tasks[task].total or len(reviews))

    if not reviews:
        console.print(f"\n  {SYM_FAIL} [bold red]No reviews found.[/bold red]")
        return

    console.print(f"\n  {SYM_OK} Downloaded: [bold green]{len(reviews):,}[/bold green] reviews")

    # Preview
    console.print()
    preview = Table(
        title="Preview (first 5 reviews)",
        border_style="bright_cyan",
        box=ROUNDED,
        title_style="bold bright_cyan",
        row_styles=["", "dim"],
    )
    preview.add_column("Steam ID", style="dim", width=20)
    preview.add_column("Author", style="bright_cyan", width=18)
    preview.add_column("Review", style="white", width=40, no_wrap=False)
    preview.add_column("Type", justify="center", width=10)
    preview.add_column("Purchase", justify="center", width=10)
    preview.add_column("Language", justify="center", style="bright_magenta", width=10)
    preview.add_column("Playtime", justify="right", style="bold", width=10)
    preview.add_column("Date", justify="center", style="dim", width=12)

    for r in reviews[:5]:
        dt = datetime.fromtimestamp(r["date"]).strftime("%Y-%m-%d")
        text = r["review"][:120] + ("..." if len(r["review"]) > 120 else "")
        type_str = "[bold green]positive[/bold green]" if r["review_type"] == "positive" else "[bold red]negative[/bold red]"
        preview.add_row(r["steam_id"], r["author"], text, type_str, r["purchase_type"], r["language"], r["playtime"], dt)

    console.print(preview)

    # Export
    safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in game_name)
    base_name = f"reviews_{safe_name}_{lang}"
    fields = ["steam_id", "author", "review", "review_type", "purchase_type", "language", "playtime", "date"]

    console.print()
    fmt = Prompt.ask(
        "[bold]Save format[/bold]",
        choices=["csv", "excel", "both", "skip"],
        default="csv",
    )

    if fmt == "skip":
        return

    rows = []
    for r in reviews:
        row = dict(r)
        row["date"] = datetime.fromtimestamp(r["date"]).strftime("%Y-%m-%d %H:%M")
        rows.append(row)

    saved_files = []

    if fmt in ("csv", "both"):
        filename_csv = f"{base_name}.csv"
        with open(filename_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(rows)
        saved_files.append(filename_csv)

    if fmt in ("excel", "both"):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        filename_xlsx = f"{base_name}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Reviews"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B2838", end_color="1B2838", fill_type="solid")

        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, field in enumerate(fields, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[field])

        for col_idx, field in enumerate(fields, 1):
            if field == "review":
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 60
            else:
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

        wb.save(filename_xlsx)
        saved_files.append(filename_xlsx)

    # Summary panel
    summary_table = Table(show_header=False, box=None, padding=(0, 2), expand=False)
    summary_table.add_column("Label", style="dim")
    summary_table.add_column("Value")

    summary_table.add_row("Reviews", f"[bold]{len(reviews):,}[/bold]")
    summary_table.add_row("Language", f"[bright_cyan]{lang_label}[/bright_cyan]")
    summary_table.add_row("Format", f"[bright_magenta]{fmt}[/bright_magenta]")
    for f_name in saved_files:
        summary_table.add_row(f"{SYM_OK} Saved", f"[bold green]{f_name}[/bold green]")

    console.print()
    console.print(Panel(
        summary_table,
        title="✅ Export Complete",
        border_style="bright_green",
        box=DOUBLE,
        expand=False,
        padding=(1, 3),
    ))


# ─── Main Loop ────────────────────────────────────────────────────────────────

def _handle_game(appid: int) -> None:
    """Load and display details for a single game, optionally scrape reviews."""
    with console.status("[bright_cyan]Loading game details...[/bright_cyan]", spinner="dots"):
        details = get_app_details(appid)
        if not details:
            console.print(f"  {SYM_FAIL} [bold red]Could not load game details.[/bold red]")
            return
        meta = parse_metadata(details)
        review_summary = get_review_summary(appid)

    show_metadata(meta, review_summary)

    console.print()
    if Confirm.ask("[bold]Download reviews?[/bold]"):
        scrape_reviews(appid, meta["title"])


def main():
    _print_banner()

    while True:
        choice = show_main_menu()

        if choice == "quit":
            break

        if choice == "search":
            query = Prompt.ask("[bold]Search a game on Steam[/bold]")
            if not query.strip():
                continue

            with console.status(f"[bright_cyan]Searching '{query}'...[/bright_cyan]", spinner="dots"):
                results = search_games(query.strip())

            if results:
                with console.status("[bright_cyan]Loading details...[/bright_cyan]", spinner="dots"):
                    enrich_search_results(results)

            appid = show_search_results(results)
            if appid is None:
                continue

            _handle_game(appid)

        else:
            # Browse category
            cat_label = BROWSE_CATEGORIES[choice]
            with console.status(f"[bright_cyan]Loading {cat_label}...[/bright_cyan]", spinner="dots"):
                results = browse_category(choice)

            if results:
                with console.status("[bright_cyan]Loading details...[/bright_cyan]", spinner="dots"):
                    enrich_search_results(results)

            appid = show_search_results(results, title=cat_label)
            if appid is None:
                continue

            _handle_game(appid)

    console.print(f"\n  {SYM_OK} [bold bright_cyan]Goodbye![/bold bright_cyan]\n")
